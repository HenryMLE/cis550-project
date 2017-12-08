from bs4 import BeautifulSoup
import openpyxl
import urllib2
import pprint
import json
import re

pp = pprint.PrettyPrinter(indent=4)

phrases = set()
wb = openpyxl.load_workbook('ABBREV_clean_all.xlsx')
st = wb.active

for i in range(st.max_row):
    phrases.add(st.cell(row = i+1, column = 2).value)


def sim_score(food, phrases):
    food = food.split(' ')
    
    bestmatch = ""
    bestscore = 0

    for p in phrases:
        parts = p.split(' ')
        words = set()
        for part in parts:
            words.add(part)

        total = 0
        score = 0

        for f in food:
            if(f in words):
                score += 1
            total += 1

        score = float(float(score) / float(total))
        if(score > bestscore) :
            bestscore = score
            bestmatch = p
    
    return bestmatch


def get_recipe(url):
    req = urllib2.Request(url)
    res = urllib2.urlopen(req)
    doc = res.read()

    soup = BeautifulSoup(doc, 'html.parser')
    title = soup.find(class_="o-AssetTitle__a-Headline").text
    ingredients = soup.find_all(class_="o-Ingredients__a-ListItemText")
    directions = soup.find(class_="o-Method__m-Body").text

    parsedIngrd = []

    for i in ingredients:
        qty = 1
        nums = re.findall("\d\/\d|\d+", i.text)
        if(len(nums) != 0):
            if("/" in nums[0]):
                frac = nums[0];
                qty = float(float(frac[0])/float(frac[2]))
            else:
                qty = float(nums[0])

        if(len(nums) > 1):
            if("/" in nums[1]):
                frac = nums[1]
                qty += float(float(frac[0])/float(frac[2]))

        text = re.sub(r'[^\w\s]','',i.text)
        text = re.sub(r' ?\d+ ?', '', text)
        fd = sim_score(text, phrases)
        parsedIngrd.append({"quantity" : qty, "food" : fd})
    
    recipe = {
        "title" : title,
        "ingredients" : parsedIngrd,
        "directions" : directions
    }

    return recipe



def get_urls():
    extensions = ["123", "a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "xyz"]

    for ext in extensions:
        print("on page " + ext)
        recipes = []
        req = urllib2.Request("http://www.foodnetwork.com/recipes/food-network-kitchen/" + ext)
        res = urllib2.urlopen(req)
        doc = res.read()

        soup = BeautifulSoup(doc, 'html.parser')
        urls = soup.find_all(class_="m-PromoList__a-ListItem", )
        for url in urls:
            try:
                recipes.append(get_recipe("http:" + url.a['href']))
            except Exception as e:
                print("skipping recipe")

        fname = ext + ".json";
        with open(fname, 'w') as f:
            json.dump(recipes, f)
    

    



get_urls()
# get_recipe("http://www.foodnetwork.com/recipes/food-network-kitchen/3-ingredient-gluten-free-banana-pancakes-3363581")

                

